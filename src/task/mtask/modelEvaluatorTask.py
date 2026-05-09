# -*- coding: utf-8 -*-
"""
@Time    : 2026/4/30
@Author  : chen
@Software: PyCharm
"""

from src.base.lib import logger, Dict,Any, os
from src.base.evaluate.model import BaseEvaluator
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font


class ModelEvaluator(BaseEvaluator):

    def evaluate(self, train_metric: Dict[str, Any], oot_metric: Dict[str, Any], cv_metric: Dict[str, Any]) -> Dict[str, Any]:

        results = {
            'oot_metrics': oot_metric,
            'train_metrics': train_metric,
            'cv_metrics': cv_metric
        }

        self.save(results)
        return results

    def save(self, evaluation_results: Dict[str, Any], output_path: str = "reports/model_evaluation.xlsx"):
        """
        Save comprehensive evaluation results to Excel file.
        """

        def fmt(v):
            if v is None:
                return None
            return round(float(v), 2)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "evaluation_summary"

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        font_white_bold = Font(color="FFFFFF", bold=True)
        alignment_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        METRIC_KEYS = ['auc', 'ks', 'recall', 'precision', 'f1', 'fp_rate', 'fn_rate', 'tp_rate', 'tn_rate']

        headers = ["data_set"] + METRIC_KEYS

        rows = []

        train_metrics = evaluation_results['train_metrics']
        rows.append(["train"] + [fmt(train_metrics.get(k)) for k in METRIC_KEYS])

        oot_metrics = evaluation_results['oot_metrics']
        rows.append(["OOT"] + [fmt(oot_metrics.get(k)) for k in METRIC_KEYS])

        if evaluation_results.get('cv_metrics') is not None:
            for fold in evaluation_results['cv_metrics']['fold_details']:
                rows.append([f"cv-fold-{fold['fold_index']}"] + [fmt(fold.get(k)) for k in METRIC_KEYS])


            cv_summary = evaluation_results['cv_metrics']['summary_stats']
            cv_mean_row = ["cv-mean"]
            for key in METRIC_KEYS:
                mean_key = f"train_{key}_mean"
                cv_mean_row.append(fmt(cv_summary.get(mean_key)))
            rows.append(cv_mean_row)

            cv_std_row = ["cv-std"]
            for key in METRIC_KEYS:
                std_key = f"train_{key}_std"
                cv_std_row.append(fmt(cv_summary.get(std_key)))
            rows.append(cv_std_row)


        start_col = 2
        for i, header in enumerate(headers):
            col = start_col + i
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = font_white_bold
            cell.alignment = alignment_center
            cell.border = thin_border


        for row_idx, row_data in enumerate(rows, start=2):
            for i, value in enumerate(row_data):
                col = start_col + i
                cell_value = "" if value is None else value
                cell = ws.cell(row=row_idx, column=col, value=cell_value)
                cell.alignment = alignment_center
                cell.border = thin_border

        for i in range(len(headers)):
            col_letter = ws.cell(row=1, column=start_col + i).column_letter
            ws.column_dimensions[col_letter].width = 14  # slightly narrower for more columns

        wb.save(output_path)
        logger.info(f"Evaluation results saved to {output_path}")